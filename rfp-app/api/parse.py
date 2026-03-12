"""POST /api/parse — Execute Document Parser pipeline, return extracted questions."""
import os, json, io, re
import requests as http_requests
from http.server import BaseHTTPRequestHandler

AIRIA_BASE        = os.environ.get("AIRIA_BASE", "https://demo.api.airia.ai")
AIRIA_API_KEY     = os.environ.get("AIRIA_API_KEY", "")
PARSER_PIPELINE   = os.environ.get("PARSER_PIPELINE", "a6ee7f8a-4b8f-4dc1-9f47-2a793e482935")

HDR      = {"X-API-Key": AIRIA_API_KEY, "User-Agent": "AiriaClient/1.0"}
HDR_JSON = {**HDR, "Content-Type": "application/json"}


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(content_length))
            file_url = body.get("url", "")

            if not file_url:
                self._json(400, {"error": "No file URL provided"})
                return

            # Execute parser pipeline
            resp = http_requests.post(
                f"{AIRIA_BASE}/v1/PipelineExecution/{PARSER_PIPELINE}",
                json={"userInput": file_url},
                headers=HDR_JSON,
                timeout=300,
            )
            resp.raise_for_status()
            result_text = resp.json().get("result", "")

            # Extract download URL from markdown
            dl_match = re.search(r'\[Download Excel\]\((https?://[^\)]+)\)', result_text)
            if not dl_match:
                self._json(500, {"error": "No download link found in parser output", "raw": result_text[:1000]})
                return

            # Download and parse the Excel
            xl_resp = http_requests.get(dl_match.group(1), headers={"User-Agent": "AiriaClient/1.0"}, timeout=60)
            if xl_resp.status_code != 200:
                self._json(500, {"error": f"Failed to download Excel: {xl_resp.status_code}"})
                return

            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xl_resp.content), read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                self._json(500, {"error": "No questions found in Excel"})
                return

            questions = []
            for r in rows[1:]:
                questions.append({
                    "id": str(r[0] or ""),
                    "question": str(r[1] or ""),
                    "category": str(r[2] or "General"),
                })

            self._json(200, {"questions": questions, "count": len(questions)})

        except Exception as e:
            self._json(500, {"error": str(e)})

    def _json(self, status, data):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
